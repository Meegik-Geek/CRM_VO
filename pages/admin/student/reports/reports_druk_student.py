
from docxtpl import DocxTemplate
import os
from tkinter import Tk, filedialog
from db.connect_db import setup_database, close_database, get_setting
from openpyxl import Workbook
class DocumentPrinter:
    def __init__(self, show_success_message, show_error_message):
        """Ініціалізація принтера з повідомленнями про успіх і помилки."""
        self.conn = None
        self.cursor = None
        self.show_success_message = show_success_message
        self.show_error_message = show_error_message

    def connect_db(self):
        """Відкриває підключення до бази даних."""
        self.conn = setup_database()
        self.cursor = self.conn.cursor()

    def close_db(self):
        """Закриває підключення до бази даних."""
        if self.conn:
            close_database(self.conn)

    def _get_common_data(self):
        """Повертає загальні дані для всіх звітів."""
        return {
            "institution_name": get_setting("college_name", "Назва закладу"),
            "institution_short_name": get_setting("college_short_name", "Скорочена назва"),
            "institution_address": get_setting("college_address", "Місто"),
            "resp_secretary": get_setting("resp_secretary", "Відповідальний секретарь ПК"),
            "deputy_secretary": get_setting("deputy_secretary", "Заступник відповідального секретаря ПК"),
            "legal_counsel": get_setting("legal_counsel", "Юристконсульт"),
            "edebo_admin": get_setting("edebo_admin", "Відповідальний за ЄДЕБО"),
            "director_name": get_setting("director_name", "Директор"),
        }

    def fetch_data(self, query, params ):
        """Виконує SQL-запит і повертає дані."""
        self.cursor.execute(query, params)
        return self.cursor.fetchone()

    def print_reported_student_denne(self, specialty_name, finance_type, order_number, order_date, arrival_date, period_start, period_end, dialog):
        """Друк повідомлень для всіх студентів спеціальності у вказаний період."""
        try:
            self.connect_db()
            
            # SQL-запит для вибірки студентів у вказаний період
            query = """
            SELECT 
                pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name
            FROM 
                student s
            JOIN 
                applicant_personal_data_day pd ON s.cert_number = pd.cert_number
            JOIN 
                personal_case_day pc ON s.number_sprava_day = pc.number_sprava
            WHERE 
                pc.name_specialnosti = %s 
                AND s.finanse = %s
                AND pc.date_sprava BETWEEN %s AND %s
            ORDER BY 
                pd.last_name ASC, pd.first_name ASC;
            """
            
            # Виконання запиту
            
            self.cursor.execute(query, (specialty_name, finance_type, period_start, period_end))
            students = self.cursor.fetchall()

            if not students:
                self.show_error_message("Дані для друку не знайдено у вказаний період!")
                self.close_db()
                return

            # Формування списку студентів для шаблону
            student_list = [
                {"full_name": row[0]}  # Форматування кожного студента як словник
                for row in students
            ]

            # Контекст для шаблону
            context = {
                "students": student_list,
                "order_number": order_number,
                "order_date": order_date,
                "specialty_name": specialty_name,
                "arrival_date": arrival_date,
                "study_form_label": "денною",
            }

            # Заповнення шаблону та збереження файлу
            template_path = "templates/student/notification_template.docx"
            self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
            self.close_db()

        except Exception as e:
            self.show_error_message(f"Помилка при створенні документів: {str(e)}")
            self.close_db()
    def print_reported_student_zaoch(self, specialty_name_zaoch, finance_type, order_number, order_date, arrival_date, period_start, period_end, dialog):
            """Друк повідомлень для всіх студентів спеціальності у вказаний період."""
            try:
                self.connect_db()
                
                # SQL-запит для вибірки студентів у вказаний період
                query = """
                SELECT 
                    pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name
                FROM 
                    student s
                JOIN 
                    applicant_personal_data_evening pd ON s.cert_number = pd.cert_number
                JOIN 
                    personal_case_evening pc ON s.number_sprava_evening = pc.number_sprava
                WHERE 
                    pc.name_specialnosti = %s 
                    AND s.finanse = %s
                    AND pc.date_sprava BETWEEN %s AND %s
                ORDER BY 
                    pd.last_name ASC, pd.first_name ASC;
                """
                
                # Виконання запиту
                
                self.cursor.execute(query, (specialty_name_zaoch, finance_type, period_start, period_end))
                students = self.cursor.fetchall()

                if not students:
                    self.show_error_message("Дані для друку не знайдено у вказаний період!")
                    self.close_db()
                    return

                # Формування списку студентів для шаблону
                student_list = [
                    {"full_name": row[0]}  # Форматування кожного студента як словник
                    for row in students
                ]

                # Контекст для шаблону
                context = {
                    "students": student_list,
                    "order_number": order_number,
                    "order_date": order_date,
                    "specialty_name": specialty_name_zaoch,
                    "arrival_date": arrival_date,
                    "study_form_label": "заочною",
                }

                # Заповнення шаблону та збереження файлу
                template_path = "templates/student/notification_template.docx"
                self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
                self.close_db()

            except Exception as e:
                self.show_error_message(f"Помилка при створенні документів: {str(e)}")
                self.close_db()
    def print_reported_student_day_scor(self, specialty_name_day_scor, finance_type, order_number, order_date, arrival_date, period_start, period_end, dialog):
            """Друк повідомлень для всіх студентів спеціальності у вказаний період."""
            try:
                self.connect_db()
                
                # SQL-запит для вибірки студентів у вказаний період
                query = """
                SELECT 
                    pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name
                FROM 
                    student s
                JOIN 
                    applicant_personal_data_day pd ON s.cert_number = pd.cert_number
                JOIN 
                    personal_case_day_scor pc ON s.number_sprava_day_scor = pc.number_sprava
                WHERE 
                    pc.name_specialnosti = %s 
                    AND s.finanse = %s
                    AND pc.date_sprava BETWEEN %s AND %s
                ORDER BY 
                    pd.last_name ASC, pd.first_name ASC;
                """
                
                # Виконання запиту
                
                self.cursor.execute(query, (specialty_name_day_scor, finance_type, period_start, period_end))
                students = self.cursor.fetchall()

                if not students:
                    self.show_error_message("Дані для друку не знайдено у вказаний період!")
                    self.close_db()
                    return

                # Формування списку студентів для шаблону
                student_list = [
                    {"full_name": row[0]}  # Форматування кожного студента як словник
                    for row in students
                ]

                # Контекст для шаблону
                context = {
                    "students": student_list,
                    "order_number": order_number,
                    "order_date": order_date,
                    "specialty_name": specialty_name_day_scor,
                    "arrival_date": arrival_date,
                    "study_form_label": "денною",
                }

                # Заповнення шаблону та збереження файлу
                template_path = "templates/student/notification_template.docx"
                self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
                self.close_db()

            except Exception as e:
                self.show_error_message(f"Помилка при створенні документів: {str(e)}")
                self.close_db()
    def print_vitag_nakaz_denne(self, specialty_name, finance_type, order_number, order_date, protokol_number, protokol_date, zarah_date, period_start, period_end, dialog):
        """Друк повідомлень для всіх студентів спеціальності у вказаний період."""
        try:
            self.connect_db()
            
            # SQL-запит для вибірки студентів у вказаний період
            query = """
            SELECT 
                pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name
            FROM 
                student s
            JOIN 
                applicant_personal_data_day pd ON s.cert_number = pd.cert_number
            JOIN 
                personal_case_day pc ON s.number_sprava_day = pc.number_sprava
            WHERE 
                pc.name_specialnosti = %s 
                AND s.finanse = %s
                AND pc.date_sprava BETWEEN %s AND %s
            ORDER BY 
                pd.last_name ASC, pd.first_name ASC;
            """
            
            # Виконання запиту
            
            self.cursor.execute(query, (specialty_name, finance_type, period_start, period_end))
            students = self.cursor.fetchall()

            if not students:
                self.show_error_message("Дані для друку не знайдено у вказаний період!")
                self.close_db()
                return
            
            
            # Формування списку студентів для шаблону
            student_list = [
                {"full_name": row[0]}  # Форматування кожного студента як словник
                for row in students
            ]
            if finance_type == "Державна форма": 
                finance_type_tm = "державного (регіонального) бюджету"
            else:
                finance_type_tm = "фізичних та/або юридичних осіб"
           
            # Контекст для шаблону
            context = {
                "students": student_list,
                "order_number": order_number,
                "finance_type": finance_type_tm,
                "order_date": order_date,
                "specialty_name": specialty_name,
                "protokol_number": protokol_number,
                "protokol_date": protokol_date, 
                "zarah_date": zarah_date,
                "study_form_label": "денної",
            }

            # Заповнення шаблону та збереження файлу
            template_path = "templates/student/extract_order_template.docx"
            self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
            self.close_db()

        except Exception as e:
            self.show_error_message(f"Помилка при створенні документів: {str(e)}")
            self.close_db()


    def print_vitag_nakaz_zaoch(self, specialty_name_zaoch, finance_type,  order_number, order_date, protokol_number, protokol_date, zarah_date, period_start, period_end, dialog):
        """Друк повідомлень для всіх студентів спеціальності у вказаний період."""
        try:
            self.connect_db()
            
            # SQL-запит для вибірки студентів у вказаний період
            query = """
            SELECT 
                pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name
            FROM 
                student s
            JOIN 
                applicant_personal_data_evening pd ON s.cert_number = pd.cert_number
            JOIN 
                personal_case_evening pc ON s.number_sprava_evening = pc.number_sprava
            WHERE 
                pc.name_specialnosti = %s 
                AND s.finanse = %s
                AND pc.date_sprava BETWEEN %s AND %s
            ORDER BY 
                pd.last_name ASC, pd.first_name ASC;
            """
            
            # Виконання запиту
            
            self.cursor.execute(query, (specialty_name_zaoch, finance_type, period_start, period_end))
            students = self.cursor.fetchall()

            if not students:
                self.show_error_message("Дані для друку не знайдено у вказаний період!")
                self.close_db()
                return

            # Формування списку студентів для шаблону
            student_list = [
                {"full_name": row[0]}  # Форматування кожного студента як словник
                for row in students
            ]

            if finance_type == "Державна форма": 
                finance_type_tm = "державного (регіонального) бюджету"
            else:
                finance_type_tm = "фізичних та/або юридичних осіб"
           
            # Контекст для шаблону
            context = {
                "students": student_list,
                "order_number": order_number,
                "finance_type": finance_type_tm,
                "order_date": order_date,
                "specialty_name": specialty_name_zaoch,
                "protokol_number": protokol_number,
                "protokol_date": protokol_date, 
                "zarah_date": zarah_date,
                "study_form_label": "заочної",
            }

            # Заповнення шаблону та збереження файлу
            template_path = "templates/student/extract_order_template.docx"
            self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
            self.close_db()

        except Exception as e:
            self.show_error_message(f"Помилка при створенні документів: {str(e)}")
            self.close_db()

    def print_vitag_nakaz_day_scor(self, specialty_name_day_scor, finance_type,  order_number, order_date, protokol_number, protokol_date, zarah_date, period_start, period_end, dialog):
        """Друк повідомлень для всіх студентів спеціальності у вказаний період."""
        try:
            self.connect_db()
            
            # SQL-запит для вибірки студентів у вказаний період
            query = """
            SELECT 
                pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name
            FROM 
                student s
            JOIN 
                applicant_personal_data_day pd ON s.cert_number = pd.cert_number
            JOIN 
                personal_case_day_scor pc ON s.number_sprava_day_scor = pc.number_sprava
            WHERE 
                pc.name_specialnosti = %s 
                AND s.finanse = %s
                AND pc.date_sprava BETWEEN %s AND %s
            ORDER BY 
                pd.last_name ASC, pd.first_name ASC;
            """
            
            # Виконання запиту
            
            self.cursor.execute(query, (specialty_name_day_scor, finance_type, period_start, period_end))
            students = self.cursor.fetchall()

            if not students:
                self.show_error_message("Дані для друку не знайдено у вказаний період!")
                self.close_db()
                return

            # Формування списку студентів для шаблону
            student_list = [
                {"full_name": row[0]}  # Форматування кожного студента як словник
                for row in students
            ]

            if finance_type == "Державна форма": 
                finance_type_tm = "державного (регіонального) бюджету"
            else:
                finance_type_tm = "фізичних та/або юридичних осіб"
           
            # Контекст для шаблону
            context = {
                "students": student_list,
                "order_number": order_number,
                "finance_type": finance_type_tm,
                "order_date": order_date,
                "specialty_name": specialty_name_day_scor,
                "protokol_number": protokol_number,
                "protokol_date": protokol_date, 
                "zarah_date": zarah_date,
                "study_form_label": "денної",
            }

            # Заповнення шаблону та збереження файлу
            template_path = "templates/student/extract_order_template.docx"
            self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
            self.close_db()

        except Exception as e:
            self.show_error_message(f"Помилка при створенні документів: {str(e)}")
            self.close_db()



    def print_list_grup(self, number_group, dialog):
        """Друк списку студентів для конкретної групи."""
        try:
            self.connect_db()
            print(number_group)
            if "Д" in number_group:
                table_applicant = "applicant_personal_data_day"
                table_personal_case = "personal_case_day"
                group_column = "number_sprava_day"
            elif "З" in number_group:
                table_applicant = "applicant_personal_data_evening"
                table_personal_case = "personal_case_evening"
                group_column = "number_sprava_evening"
            else:
                self.show_error_message("Неправильний формат номера групи! Відсутній символ 'Д' або 'З'.")
                self.close_db()
                return

            # SQL-запит для отримання студентів конкретної групи
            query = f"""
                SELECT 
                    pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name,
                    s.name_specialnosti
                FROM 
                    student s
                JOIN 
                    {table_applicant} pd ON s.cert_number = pd.cert_number
                JOIN 
                    {table_personal_case} pc ON s.{group_column} = pc.number_sprava
                WHERE 
                    s.group_number = %s
                ORDER BY 
                    pd.last_name ASC, pd.first_name ASC;
            """
            
            # Виконання запиту
            self.cursor.execute(query, (number_group,))
            students = self.cursor.fetchall()

            if not students:
                self.show_error_message("Дані для друку не знайдено для цієї групи!")
                self.close_db()
                return

            # Формування контексту для шаблону
            student_list = [
                {"index": i + 1, "full_name": student[0]}
                for i, student in enumerate(students)
            ]
            context = {
                "number_group": number_group,
                "name_specialnosti": students[0][1],  # Назва спеціальності (однакова для всіх студентів)
                "students": student_list
            }

            # Шлях до шаблону
            template_path = "templates/student/list_grup_template.docx"
            
            # Заповнюємо шаблон і зберігаємо файл
            self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
            self.close_db()

        except Exception as e:
            self.show_error_message(f"Помилка при створенні документів: {str(e)}")
            self.close_db()




    def print_list_grup_roz(self, number_group, dialog):
        """Друк списку студентів для конкретної групи."""
        try:
            self.connect_db()
            
            if "Д" in number_group:
                table_applicant = "applicant_personal_data_day"
                table_personal_case = "personal_case_day"
                group_column = "number_sprava_day"
            elif "З" in number_group:
                table_applicant = "applicant_personal_data_evening"
                table_personal_case = "personal_case_evening"
                group_column = "number_sprava_evening"
            else:
                self.show_error_message("Неправильний формат номера групи! Відсутній символ 'Д' або 'З'.")
                self.close_db()
                return

            # SQL-запит для отримання студентів конкретної групи
            query = f"""
                 SELECT 
                    pd.last_name || ' ' || pd.first_name || ' ' || COALESCE(pd.middle_name, '') AS full_name,
                    pd.date_birth,
                    pd.address,
                    s.finanse,
                    s.name_specialnosti
                FROM 
                    student s
                JOIN 
                    {table_applicant} pd ON s.cert_number = pd.cert_number
                JOIN 
                    {table_personal_case} pc ON s.{group_column} = pc.number_sprava
                WHERE 
                    s.group_number = %s
                ORDER BY 
                    pd.last_name ASC, pd.first_name ASC;
            """
            
            # Виконання запиту
            self.cursor.execute(query, (number_group,))
            students = self.cursor.fetchall()

            if not students:
                self.show_error_message("Дані для друку не знайдено для цієї групи!")
                self.close_db()
                return

            # Статичні дані
            context = {
                "number_group": number_group,
                "name_specialnosti": students[0][4]  # Назва спеціальності
            }

            # Динамічні дані студентів
            student_list = [
                {
                    "index": i + 1,
                    "full_name": student[0],
                    "date_birth": student[1],
                    "address": student[2],
                    "finance": student[3]
                }
                for i, student in enumerate(students)
            ]

            # Об'єднання контексту
            context["students"] = student_list



            template_path = "templates/student/list_grup_roz_template.docx"
            
            # Заповнюємо шаблон і зберігаємо файл
            self.fill_and_save_template(template_path, context, dialog, "Документ успішно збережено!")
            self.close_db()

        except Exception as e:
            self.show_error_message(f"Помилка при створенні документів: {str(e)}")
            print(f"Помилка при створенні документів: {str(e)}")
            self.close_db()



    def print_export_date_vstupnik(self):
        """Експортує дані з бази в Excel з поділом за формами навчання та групами."""
        try:
            self.connect_db()

            # Загальний запит для вибірки даних
            query = """
            SELECT 
                os.number_sprava,
                os.date_sprava,
                s.finanse,
                pd.last_name,
                pd.first_name,
                pd.middle_name,
                pd.pip,
                pd.phone,
                pd.citizenship,
                pd.passport_number,
                pd.issued_by,
                pd.issue_date,
                pd.id_code,
                pd.cert_number,
                pd.cert_issue_date,
                pd.address,
                pd.father_first_name,
                pd.father_last_name,
                pd.father_middle_name,
                pd.father_job,
                pd.father_phone,
                pd.mother_first_name,
                pd.mother_last_name,
                pd.mother_middle_name,
                pd.mother_phone,
                pd.mother_job,
                pd.hostel_need,
                pd.gender,
                pd.algebra,
                pd.geometry,
                pd.ukr_language,
                pd.ukr_literature,
                pd.school_name,
                pd.date_birth,
                STRING_AGG(pv.kod_pilgi::TEXT, ', ') AS kod_pilgi,
                STRING_AGG(pv.document_pilgi, ', ') AS document_pilgi,
                s.group_number
            FROM 
                student s
            JOIN 
                {applicant_table} pd ON s.cert_number = pd.cert_number
            JOIN 
                {case_table} os ON s.{group_column} = os.number_sprava
            LEFT JOIN 
                {benefits_table} pv ON pv.cert_number = pd.cert_number
            GROUP BY 
                os.number_sprava, os.date_sprava, s.finanse, pd.last_name, pd.first_name,  
                pd.middle_name, pd.pip, pd.phone, pd.citizenship, pd.passport_number, pd.issued_by, 
                pd.issue_date, pd.id_code, pd.cert_number, pd.cert_issue_date, pd.address, pd.father_first_name, pd.father_last_name, 
                pd.father_middle_name, pd.father_job, pd.father_phone, pd.mother_first_name, 
                pd.mother_last_name, pd.mother_middle_name, pd.mother_phone, pd.mother_job, 
                pd.hostel_need, pd.gender, pd.algebra, pd.geometry, pd.ukr_language, 
                pd.ukr_literature, pd.school_name, pd.cert_issue_date, pd.date_birth, s.group_number
            ORDER BY 
                s.group_number, pd.last_name, pd.first_name;
            """

            # Вибірка даних для денної форми
            query_denna = query.format(
                applicant_table="applicant_personal_data_day",
                case_table="personal_case_day",
                group_column="number_sprava_day",
                benefits_table="applicant_benefits_day"
            )
            self.cursor.execute(query_denna)
            data_denna = self.cursor.fetchall()

            # Вибірка даних для заочної форми
            query_zaoch = query.format(
                applicant_table="applicant_personal_data_evening",
                case_table="personal_case_evening",
                group_column="number_sprava_evening",
                benefits_table="applicant_benefits_evening"
            )
            self.cursor.execute(query_zaoch)
            data_zaoch = self.cursor.fetchall()

            # Перевірка, чи є дані
            if not data_denna and not data_zaoch:
                self.show_error_message("Немає даних для експорту.")
                self.close_db()
                return

            # Виклик діалогу для вибору шляху збереження файлу
            root = Tk()
            root.withdraw()
            file_path = filedialog.asksaveasfilename(
                title="Зберегти файл як...",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )
            if not file_path:
                self.show_success_message("Експорт скасовано користувачем.")
                self.close_db()
                return

            # Створення Excel-файлу
            workbook = Workbook()
            headers = [
                "Номер справи", "Дата подачі справи", "Фінансування", "Прізвище", "Ім'я", 
                "По батькові", "ПІП", "Телефон", "Громадянство", "Номер паспорта", "Ким виданий",
                "Дата видачі паспорта", "Ідентифікаційний код", "Свідоцтво про освіту", "Дата видачі свідоцтва", "Адреса", "Ім'я батька",
                "Прізвище батька", "По батькові батька", "Робота батька", "Телефон батька",
                "Ім'я матері", "Прізвище матері", "По батькові матері", "Телефон матері", "Робота матері",
                "Потреба в гуртожитку", "Стать", "Алгебра", "Геометрія", "Українська мова",
                "Українська література", "Назва школи",  "Дата народження",
                "Коди пільг", "Документи про пільги"
            ]

            # Додавання даних до окремих листів для денної форми
            if data_denna:
                data_by_group_denna = {}
                for row in data_denna:
                    group_number = row[-1]
                    if group_number not in data_by_group_denna:
                        data_by_group_denna[group_number] = []
                    data_by_group_denna[group_number].append(row[:-1])

                for group, data in data_by_group_denna.items():
                    sheet = workbook.create_sheet(title=f"Денна група {group}")
                    sheet.append(headers)
                    for row in data:
                        sheet.append(row)

            # Додавання даних до окремих листів для заочної форми
            if data_zaoch:
                data_by_group_zaoch = {}
                for row in data_zaoch:
                    group_number = row[-1]
                    if group_number not in data_by_group_zaoch:
                        data_by_group_zaoch[group_number] = []
                    data_by_group_zaoch[group_number].append(row[:-1])

                for group, data in data_by_group_zaoch.items():
                    sheet = workbook.create_sheet(title=f"Заочна група {group}")
                    sheet.append(headers)
                    for row in data:
                        sheet.append(row)

            # Видалення стандартного листа, якщо він залишився порожнім
            if "Sheet" in workbook.sheetnames and not workbook["Sheet"].max_row:
                del workbook["Sheet"]

            # Збереження файлу
            workbook.save(file_path)
            self.show_success_message(f"Дані успішно експортовано до {os.path.basename(file_path)}")

        except Exception as e:
            self.show_error_message(f"Помилка при експорті даних: {str(e)}")
        finally:
            self.close_db()

   
    def fill_and_save_template(self, template_path, data, dialog, success_message):
        """Заповнює шаблон DOCX і зберігає документ із вибором місця збереження."""
        try:
            common = self._get_common_data()
            common.update(data)
            doc = DocxTemplate(template_path)
            doc.render(common)

            # Показуємо діалог для вибору місця збереження
            root = Tk()
            root.withdraw()
            save_path = filedialog.asksaveasfilename(
                defaultextension=".docx",
                filetypes=[("Word Documents", "*.docx")],
                initialfile="Звіт",
                title="Оберіть шлях для збереження документа"
            )
            root.destroy()

            if save_path:
                doc.save(save_path)
                dialog.accept()  # Закриваємо діалог після успішного друку
                self.show_success_message(success_message)
            else:
                self.show_error_message("Збереження скасовано користувачем.")

        except Exception as e:
            self.show_error_message(f"Помилка при створенні документів: {str(e)}")
            print(f"Помилка при створенні документів: {str(e)}")

